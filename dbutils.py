"""
File: dbutils.py
Author: SluG
Created: February 28, 2024
Description: Module that provides tools to manage the different databases
"""

import pandas as pd
import os
from loguru import logger
import openpyxl

def excel_file_exists(file_path):
    """
    Check if an Excel file exists at the given file path.

    Parameters:
    - file_path (str): The path to the Excel file.

    Returns:
    - bool: True if the file exists, False otherwise.
    """
    logger.info("Start excel_file_exists")
    logger.info("Parameter file_path: " + file_path)
    return os.path.isfile(file_path)


def create_empty_excel_database(file_name):
    """
    Create an empty Excel database at the specified file path.

    Parameters:
    - file_path (str): The path where the Excel file will be saved.
    """
    logger.info("Start create_empty_excel module")
    logger.info("Parameter file_name: " + file_name)
    file_path = file_name

    if excel_file_exists(file_name):        
        logger.info("The Excel file " + file_name + " exists, so we don't create it again.")
    else:
        logger.info("The Excel file " + file_name + " does not exist, so we create it.")
        # We create the workbook
        workbook = openpyxl.Workbook()
        # We get the active sheet
        sheet = workbook.active
        # We save the file
        workbook.save(file_name)
        # We change the name of the sheets and create the headers
        if(file_name == "emails.xlsx"):
            # We create the structure of the emails excel file
            # Tab Dashboard
            new_tab_name = 'Dashboard'
            cell_data = {'A1': 'TotalNumber', 'B1': 'DomainsNumber'}
            create_and_write_to_excel(file_name, new_tab_name, cell_data)
            # Tab emails
            new_tab_name = 'Emails'
            cell_data = {'A1': 'Email', 'B1': 'Domain', 'C1': 'DateAdded'}
            create_and_write_to_excel(file_name, new_tab_name, cell_data)
            # TODO: Delete "Sheet" tab of the emails excel file
        else:
            # We create the structure of the links excel file
            # Tab Dashboard
            new_tab_name = 'Dashboard'
            cell_data = {'A1': 'TotalNumber'}
            create_and_write_to_excel(file_name, new_tab_name, cell_data)        
            # Tab Links
            new_tab_name = 'Links'
            cell_data = {'A1': 'Link', 'B1': 'DateAdded', 'C1': 'Processed', 'D1': 'DateProcessed'}
            create_and_write_to_excel(file_name, new_tab_name, cell_data)
            # TODO: Delete "Sheet" tab of the links excel file

        logger.info("Excel file created successfuly: " + file_name)

def initializeDBs():
    """
    Initialize the databases

    Parameters:
    
    """
    logger.info("Start initializeDBs")
    # Definition of the databases names
    linksDBname = "links.xlsx"
    emailsDBname = "emails.xlsx"
    # Check that databases exist. If not, is the first time we run the agent, so we create them
    create_empty_excel_database(linksDBname)
    logger.info("Links DB initialized")
    create_empty_excel_database(emailsDBname)
    logger.info("Emails DB initialized")

def write_to_excel(file_path, sheet_name, cell_address, text_to_write):
    """
    Write text to a specific cell in an Excel file.

    Parameters:
    - file_path (str): The path to the Excel file.
    - sheet_name (str): The name of the sheet in the Excel file.
    - cell_address (str): The address of the cell (e.g., 'A1', 'B2').
    - text_to_write (str): The text to write to the cell.
    """
    logger.info("Start write_to_excel module")
    logger.info("Parameter file_path: " + file_path)
    logger.info("Parameter sheet_name: " + sheet_name)
    logger.info("Parameter cell_address: " + cell_address)
    logger.info("Parameter text_to_write: " + text_to_write)
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    # Select the sheet
    sheet = workbook[sheet_name]
    # Write text to the specified cell
    sheet[cell_address] = text_to_write
    # Save the changes
    workbook.save(file_path)

def rename_excel_sheet(file_path, current_sheet_name, new_sheet_name):
    """
    Rename a sheet in an Excel file.

    Parameters:
    - file_path (str): The path to the Excel file.
    - current_sheet_name (str): The current name of the sheet to be renamed.
    - new_sheet_name (str): The new name for the sheet.
    """
    logger.info("Start rename_excel_sheet module")
    logger.info("Parameter file_path: " + file_path)
    logger.info("Parameter current_sheet_name: " + current_sheet_name)
    logger.info("Parameter new_sheet_name: " + new_sheet_name)
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    # Check if the sheet exists
    if current_sheet_name in workbook.sheetnames:
        # Get the sheet
        sheet = workbook[current_sheet_name]
        # Rename the sheet
        sheet.title = new_sheet_name
        # Save the changes
        workbook.save(file_path)
        logger.info("The sheet '{current_sheet_name}' has been renamed to '{new_sheet_name}'.")
    else:
        logger.info("The sheet '{current_sheet_name}' does not exist in the Excel file.")

def create_and_write_to_excel(file_path, new_tab_name, cell_data):
    """
    Create a new tab in an Excel file and write text to multiple cells.

    Parameters:
    - file_path (str): The path to the Excel file.
    - new_tab_name (str): The name for the new tab.
    - cell_data (dict): A dictionary where keys are cell addresses (e.g., 'A1', 'B2') 
                       and values are the text to write to each cell.
    """
    logger.info("Start create_and_write_to_excel module")
    logger.info("Parameter file_path: " + file_path)
    logger.info("Parameter new_tab_name: " + new_tab_name)
    logger.info("Parameter cell_data: " + cell_data)    
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    # Create a new worksheet
    new_tab = workbook.create_sheet(title=new_tab_name)
    # Write text to multiple cells in the new worksheet
    for cell_address, text_to_write in cell_data.items():
        new_tab[cell_address] = text_to_write
    # Save the changes
    workbook.save(file_path)
    logger.info("The new tab '{new_tab_name}' has been created, and text has been written to the specified cells.")
